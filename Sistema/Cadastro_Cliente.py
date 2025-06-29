import customtkinter as ctk
from tkinter import messagebox
from PIL import Image
import os, re
from openpyxl import Workbook, load_workbook

ARQUIVO_EXCEL = r"C:\Users\Hemerson\Desktop\Projeto Olympus\Site\clientes.xlsx"

# ---------- funções utilitárias ----------
def validar_email(email):
    return re.match(r'^[\w\.-]+@[\w\.-]+\.\w{2,}$', email)

def gerar_proximo_codigo():
    if not os.path.exists(ARQUIVO_EXCEL):
        return 1
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        sheet = wb["Clientes"] if "Clientes" in wb.sheetnames else wb.active
        codigos = []
        for row in sheet.iter_rows(min_row=2, max_col=1):
            val = row[0].value
            if val is None:
                continue
            try:
                codigos.append(int(str(val)))
            except ValueError:
                pass
        return max(codigos) + 1 if codigos else 1
    except Exception:
        return 1

def preencher_codigo_automatico():
    entry_codigo.configure(state="normal")
    entry_codigo.delete(0, ctk.END)
    entry_codigo.insert(0, str(gerar_proximo_codigo()))
    entry_codigo.configure(state="disabled")

def salvar_dados():
    dados = {
        "codigo": entry_codigo.get(),
        "nome": entry_nome.get(),
        "cpf": entry_cpf.get(),
        "estado_civil": opcao_estado_civil.get(),
        "nascimento": entry_data_nasc.get(),
        "sexo": opcao_sexo.get(),
        "celular": entry_celular.get(),
        "email": entry_email.get(),
        "peso": entry_peso.get(),
        "altura": entry_altura.get(),
        "objetivo": opcao_objetivo.get(),
        "observacoes": entry_obs.get()
    }

    if not dados["nome"] or not dados["cpf"]:
        messagebox.showwarning("Atenção", "Preencha os campos Nome e CPF!")
        return
    if not validar_email(dados["email"]):
        messagebox.showwarning("Atenção", "Email inválido!")
        return

    # cria Excel se não existir
    if not os.path.exists(ARQUIVO_EXCEL):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Clientes"
        sheet.append([
            "Código", "Nome Completo", "CPF", "Estado Civil", "Data Nascimento",
            "Sexo", "Celular", "Email", "Peso", "Altura", "Objetivo", "Observações"
        ])
    else:
        wb = load_workbook(ARQUIVO_EXCEL)
        sheet = wb["Clientes"] if "Clientes" in wb.sheetnames else wb.active

    sheet.append([
        int(dados["codigo"]), dados["nome"], dados["cpf"], dados["estado_civil"],
        dados["nascimento"], dados["sexo"], dados["celular"], dados["email"],
        dados["peso"], dados["altura"], dados["objetivo"], dados["observacoes"]
    ])
    wb.save(ARQUIVO_EXCEL)
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")
    limpar_campos()
    preencher_codigo_automatico()

def limpar_campos():
    for e in (entry_nome, entry_cpf, entry_data_nasc, entry_celular,
              entry_email, entry_peso, entry_altura, entry_obs):
        e.delete(0, ctk.END)
    opcao_estado_civil.set("Solteiro(a)")
    opcao_sexo.set("Masculino")
    opcao_objetivo.set("Emagrecer")

# ---------- UI ----------
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

janela = ctk.CTk()
janela.title("Cadastro de Cliente")
janela.geometry("700x850")

frame_principal = ctk.CTkFrame(janela, corner_radius=10)
frame_principal.pack(pady=20, padx=20, fill="both", expand=True)

# topo (logo + título)
try:
    img = Image.open(r"C:\Users\Hemerson\Desktop\Projeto Olympus\Sistema\images\Logo.png")
    logo = ctk.CTkImage(light_image=img, size=(120, 120))
    ctk.CTkLabel(frame_principal, image=logo, text="").pack(pady=5)
except FileNotFoundError:
    ctk.CTkLabel(frame_principal, text="(Logo não encontrada)").pack(pady=5)

ctk.CTkLabel(frame_principal, text="Cadastro de Clientes",
             font=ctk.CTkFont(size=20, weight="bold")).pack(pady=5)

# frame com rolagem
scroll = ctk.CTkScrollableFrame(frame_principal, width=640, height=650)
scroll.pack(fill="both", expand=True, padx=10, pady=10)

def linha(row, texto, widget):
    ctk.CTkLabel(scroll, text=f"{texto}:", anchor="e", width=180)\
        .grid(row=row, column=0, sticky="e", padx=10, pady=7)
    widget.grid(row=row, column=1, sticky="w", padx=10, pady=7)

# campos do formulário
entry_codigo = ctk.CTkEntry(scroll, width=300, state="disabled")
linha(0, "Código", entry_codigo)

entry_nome = ctk.CTkEntry(scroll, width=300, placeholder_text="Digite o nome completo")
linha(1, "Nome completo", entry_nome)

entry_cpf = ctk.CTkEntry(scroll, width=300, placeholder_text="Digite o CPF")
linha(2, "CPF", entry_cpf)

entry_data_nasc = ctk.CTkEntry(scroll, width=300, placeholder_text="dd/mm/aaaa")
linha(3, "Data de nascimento", entry_data_nasc)

opcao_estado_civil = ctk.CTkOptionMenu(scroll,
    values=["Solteiro(a)", "Casado(a)", "Divorciado(a)", "Viúvo(a)"])
opcao_estado_civil.set("Solteiro(a)")
linha(4, "Estado civil", opcao_estado_civil)

opcao_sexo = ctk.CTkOptionMenu(scroll, values=["Masculino", "Feminino"])
opcao_sexo.set("Masculino")
linha(5, "Sexo", opcao_sexo)

entry_celular = ctk.CTkEntry(scroll, width=300, placeholder_text="Digite o celular")
linha(6, "Celular", entry_celular)

entry_email = ctk.CTkEntry(scroll, width=300, placeholder_text="Digite o e-mail")
linha(7, "Email", entry_email)

entry_peso = ctk.CTkEntry(scroll, width=300, placeholder_text="Ex: 70.5")
linha(8, "Peso (kg)", entry_peso)

entry_altura = ctk.CTkEntry(scroll, width=300, placeholder_text="Ex: 175")
linha(9, "Altura (cm)", entry_altura)

opcao_objetivo = ctk.CTkOptionMenu(scroll,
    values=["Emagrecer", "Ganhar Massa", "Condicionamento", "Manutenção", "Saúde geral"])
opcao_objetivo.set("Emagrecer")
linha(10, "Objetivo", opcao_objetivo)

entry_obs = ctk.CTkEntry(scroll, width=300,
                         placeholder_text="Ex: lesões, preferências alimentares...")
linha(11, "Observações", entry_obs)

# botões
btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
btn_frame.grid(row=12, column=0, columnspan=2, pady=20)

ctk.CTkButton(btn_frame, text="Salvar", command=salvar_dados,
              width=150).grid(row=0, column=0, padx=10)

ctk.CTkButton(btn_frame, text="Limpar", command=limpar_campos,
              fg_color="#f44336", hover_color="#e53935",
              width=150).grid(row=0, column=1, padx=10)

ctk.CTkButton(btn_frame, text="Atualizar Código", command=preencher_codigo_automatico,
              fg_color="#4CAF50", hover_color="#45a049",
              width=150).grid(row=0, column=2, padx=10)

# inicializa primeiro código
preencher_codigo_automatico()
janela.mainloop()
