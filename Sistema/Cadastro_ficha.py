import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

ARQUIVO_TREINO = r"C:\Users\Hemerson\Desktop\Projeto Olympus\Base_dados\Padroes.xlsx"

# ----------------- Funções -----------------
def gerar_proximo_id():
    if not os.path.exists(ARQUIVO_TREINO):
        return 1
    try:
        wb = load_workbook(ARQUIVO_TREINO)
        sheet = wb.active
        ids = [int(row[0].value) for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
        return max(ids) + 1 if ids else 1
    except:
        return 1

def preencher_id_automatico():
    entry_id.configure(state="normal")
    entry_id.delete(0, ctk.END)
    entry_id.insert(0, str(gerar_proximo_id()))
    entry_id.configure(state="disabled")

def salvar_treino():
    dados = {
        "id": entry_id.get(),
        "nivel": opcao_nivel.get(),
        "tipo": entry_tipo.get(),
        "cardio": entry_cardio.get(),
        "frequencia": entry_frequencia.get(),
        "refeicoes": entry_refeicoes.get()
    }

    if not dados["tipo"] or not dados["cardio"]:
        messagebox.showwarning("Atenção", "Preencha os campos de tipo e cardio.")
        return

    if not os.path.exists(ARQUIVO_TREINO):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Treinos"
        sheet.append(["ID", "Nível", "Tipo de Treino", "Cardio", "Frequência na Semana", "Refeições"])
    else:
        wb = load_workbook(ARQUIVO_TREINO)
        sheet = wb["Treinos"] if "Treinos" in wb.sheetnames else wb.active

    sheet.append([
        int(dados["id"]), dados["nivel"], dados["tipo"],
        dados["cardio"], dados["frequencia"], dados["refeicoes"]
    ])
    wb.save(ARQUIVO_TREINO)
    messagebox.showinfo("Sucesso", "Treino salvo com sucesso!")
    limpar_treino()
    preencher_id_automatico()

def limpar_treino():
    for e in (entry_tipo, entry_cardio, entry_frequencia, entry_refeicoes):
        e.delete(0, ctk.END)
    opcao_nivel.set("Iniciante")

# ----------------- UI -----------------
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

janela = ctk.CTk()
janela.title("Cadastro de Treinos")
janela.geometry("700x600")

frame = ctk.CTkFrame(janela, corner_radius=10)
frame.pack(pady=20, padx=20, fill="both", expand=True)

ctk.CTkLabel(frame, text="Cadastro de Treinos",
             font=ctk.CTkFont(size=20, weight="bold")).pack(pady=10)

scroll = ctk.CTkScrollableFrame(frame, width=640, height=400)
scroll.pack(padx=10, pady=10, fill="both", expand=True)

def linha(row, texto, widget):
    ctk.CTkLabel(scroll, text=f"{texto}:", anchor="e", width=180)\
        .grid(row=row, column=0, sticky="e", padx=10, pady=7)
    widget.grid(row=row, column=1, sticky="w", padx=10, pady=7)

entry_id = ctk.CTkEntry(scroll, width=300, state="disabled")
linha(0, "ID", entry_id)

opcao_nivel = ctk.CTkOptionMenu(scroll, values=["Iniciante", "Intermediário", "Avançado"])
opcao_nivel.set("Iniciante")
linha(1, "Nível", opcao_nivel)

entry_tipo = ctk.CTkEntry(scroll, width=300, placeholder_text="Ex: Musculação, Funcional")
linha(2, "Tipo de Treino", entry_tipo)

entry_cardio = ctk.CTkEntry(scroll, width=300, placeholder_text="Ex: Corrida, Bicicleta")
linha(3, "Cardio", entry_cardio)

entry_frequencia = ctk.CTkEntry(scroll, width=300, placeholder_text="Ex: 3x por semana")
linha(4, "Frequência na Semana", entry_frequencia)

entry_refeicoes = ctk.CTkEntry(scroll, width=300, placeholder_text="Ex: 5 refeições/dia")
linha(5, "Refeições", entry_refeicoes)

btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
btn_frame.grid(row=6, column=0, columnspan=2, pady=20)

ctk.CTkButton(btn_frame, text="Salvar", command=salvar_treino, width=150).grid(row=0, column=0, padx=10)
ctk.CTkButton(btn_frame, text="Limpar", command=limpar_treino, fg_color="#f44336",
              hover_color="#e53935", width=150).grid(row=0, column=1, padx=10)
ctk.CTkButton(btn_frame, text="Atualizar ID", command=preencher_id_automatico,
              fg_color="#4CAF50", hover_color="#45a049", width=150).grid(row=0, column=2, padx=10)

preencher_id_automatico()
janela.mainloop()
