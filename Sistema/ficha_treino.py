
import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

ARQUIVO_EXCEL = "fichas_treino.xlsx"

def gerar_proximo_id():
    if not os.path.exists(ARQUIVO_EXCEL):
        return 1
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        sheet = wb["Fichas"]
        ids = [row[0].value for row in sheet.iter_rows(min_row=2) if isinstance(row[0].value, int)]
        return max(ids) + 1 if ids else 1
    except Exception:
        return 1

def preencher_id_automatico():
    proximo = gerar_proximo_id()
    entry_id.configure(state="normal")
    entry_id.delete(0, ctk.END)
    entry_id.insert(0, str(proximo))
    entry_id.configure(state="disabled")

def salvar_ficha():
    id_ficha = entry_id.get()
    nome_cliente = entry_nome.get()
    objetivo = entry_objetivo.get()
    exercicios = text_exercicios.get("1.0", "end").strip()

    if not nome_cliente or not objetivo or not exercicios:
        messagebox.showwarning("Atenção", "Preencha todos os campos!")
        return

    if not os.path.exists(ARQUIVO_EXCEL):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Fichas"
        sheet.append(["ID", "Nome do Cliente", "Objetivo", "Exercícios"])
    else:
        wb = load_workbook(ARQUIVO_EXCEL)
        sheet = wb["Fichas"]

    sheet.append([int(id_ficha), nome_cliente, objetivo, exercicios])
    wb.save(ARQUIVO_EXCEL)
    messagebox.showinfo("Sucesso", "Ficha salva com sucesso!")
    limpar_campos()
    preencher_id_automatico()

def limpar_campos():
    entry_nome.delete(0, ctk.END)
    entry_objetivo.delete(0, ctk.END)
    text_exercicios.delete("1.0", "end")

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")

janela = ctk.CTk()
janela.title("Cadastro de Ficha de Treino")
janela.geometry("650x700")

frame = ctk.CTkFrame(janela, corner_radius=10)
frame.pack(pady=20, padx=20, fill="both", expand=True)

ctk.CTkLabel(frame, text="Ficha de Treino", font=ctk.CTkFont(size=22, weight="bold")).pack(pady=15)

form = ctk.CTkFrame(frame)
form.pack(pady=10, padx=20)

def criar_linha(row, texto, widget):
    ctk.CTkLabel(form, text=texto + ":", anchor="e", width=160).grid(row=row, column=0, sticky="e", padx=10, pady=7)
    widget.grid(row=row, column=1, sticky="w", padx=10, pady=7)

entry_id = ctk.CTkEntry(form, width=300)
entry_id.configure(state="disabled")
criar_linha(0, "ID", entry_id)

entry_nome = ctk.CTkEntry(form, width=300, placeholder_text="Nome do Cliente")
criar_linha(1, "Nome do Cliente", entry_nome)

entry_objetivo = ctk.CTkEntry(form, width=300, placeholder_text="Ex: Hipertrofia")
criar_linha(2, "Objetivo", entry_objetivo)

ctk.CTkLabel(form, text="Exercícios:", anchor="ne", width=160).grid(row=3, column=0, sticky="ne", padx=10, pady=7)
text_exercicios = ctk.CTkTextbox(form, width=300, height=200)
text_exercicios.grid(row=3, column=1, sticky="w", padx=10, pady=7)

botoes = ctk.CTkFrame(frame, fg_color="transparent")
botoes.pack(pady=15)

btn_salvar = ctk.CTkButton(botoes, text="Salvar Ficha", command=salvar_ficha, width=150)
btn_salvar.grid(row=0, column=0, padx=10)

btn_limpar = ctk.CTkButton(botoes, text="Limpar", command=limpar_campos, fg_color="#f44336", hover_color="#e53935", width=150)
btn_limpar.grid(row=0, column=1, padx=10)

preencher_id_automatico()
janela.mainloop()
