from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

ARQUIVO = 'clientes.xlsx'

# Cria a planilha com cabeçalho, se não existir
if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Código", "Nome", "CPF", "Nascimento", "Estado Civil", "Sexo",
        "Celular", "Email", "Peso", "Altura", "Objetivo", "Observações"
    ])
    wb.save(ARQUIVO)

@app.route('/cadastrar', methods=['POST'])
def cadastrar():
    dados = request.get_json()

    if not dados.get("nome") or not dados.get("email"):
        return jsonify({"erro": "Nome e e-mail são obrigatórios"}), 400

    wb = load_workbook(ARQUIVO)
    ws = wb.active

    # Gera código auto-incremental formatado com zeros à esquerda
    if ws.max_row == 1:  # só o cabeçalho
        novo_codigo = "0001"
    else:
        ultimo_codigo = ws.cell(row=ws.max_row, column=1).value
        try:
            novo_codigo = f"{int(ultimo_codigo) + 1:04d}"
        except ValueError:
            novo_codigo = "0001"

    # Adiciona os dados na planilha
    ws.append([
        novo_codigo,
        dados.get("nome"),
        dados.get("cpf"),
        dados.get("nascimento"),
        dados.get("estado_civil"),
        dados.get("sexo"),
        dados.get("celular"),
        dados.get("email"),
        dados.get("peso"),
        dados.get("altura"),
        dados.get("objetivo"),
        dados.get("observacoes")
    ])
    wb.save(ARQUIVO)

    return jsonify({
        "mensagem": "Cadastro salvo com sucesso!",
        "codigo": novo_codigo
    }), 200

if __name__ == '__main__':
    app.run(debug=True)

